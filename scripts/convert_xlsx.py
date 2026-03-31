# -*- coding: utf-8 -*-
import openpyxl, sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX = r'C:/Users/kjinh/Downloads/할렐루양 국문 데이터.xlsx'
BASE = r'C:/Users/kjinh/OneDrive/Desktop/edan/lib/features'

wb = openpyxl.load_workbook(XLSX, data_only=True)

def esc(s):
    if not s:
        return ''
    return str(s).replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n").replace("\r", "")


# ===== 1. Bible data =====
ws = wb['국문 성경']
chapters = {}
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[7] and row[5] and row[6]:
        ch = int(float(str(row[5])))
        verse = int(float(str(row[6])))
        text = esc(row[7])
        if ch not in chapters:
            chapters[ch] = []
        chapters[ch].append((verse, text))

os.makedirs(f'{BASE}/bible/data', exist_ok=True)
with open(f'{BASE}/bible/data/matthew_bible_data.dart', 'w', encoding='utf-8') as f:
    f.write("import '../models/bible_book.dart';\n\n")
    f.write("/// 마태복음 전체 성경 데이터 (개역한글) - xlsx에서 자동 생성\n")
    f.write("class MatthewBibleData {\n")
    f.write("  static const Map<int, List<BibleVerse>> chapters = {\n")
    for ch in sorted(chapters.keys()):
        f.write(f"    {ch}: [\n")
        for v, t in chapters[ch]:
            f.write(f"      BibleVerse(verse: {v}, text: '{t}'),\n")
        f.write("    ],\n")
    f.write("  };\n\n")
    f.write("  static BibleChapter? getChapter(int chapter) {\n")
    f.write("    final verses = chapters[chapter];\n")
    f.write("    if (verses == null) return null;\n")
    f.write("    return BibleChapter(bookName: '마태복음', chapter: chapter, verses: verses);\n")
    f.write("  }\n\n")
    f.write("  static bool hasChapter(int chapter) => chapters.containsKey(chapter);\n")
    f.write("}\n")
print(f"Bible: {len(chapters)} chapters, {sum(len(v) for v in chapters.values())} verses")


# ===== 2. Quiz data =====
ws2 = wb['국문 퀴즈_마태복음']
quizzes = {}
for row in ws2.iter_rows(min_row=3, values_only=True):
    if not (row[5] and row[6] and row[7]):
        continue
    unit_raw = row[1]
    if unit_raw is None:
        continue
    unit = int(float(str(unit_raw)))
    ch = int(float(str(row[3]))) if row[3] else 0
    order = int(float(str(row[4]))) if row[4] else 0
    question = esc(row[5])
    options_raw = str(row[6])
    answer = str(row[7]).strip()

    options = [o.strip() for o in options_raw.split('/')]
    correct_idx = 0
    for i, opt in enumerate(options):
        if opt.strip() == answer:
            correct_idx = i
            break

    if unit not in quizzes:
        quizzes[unit] = {'chapter': ch, 'items': []}
    quizzes[unit]['items'].append((order, question, options, correct_idx))

os.makedirs(f'{BASE}/study/data', exist_ok=True)
with open(f'{BASE}/study/data/matthew_quiz_data.dart', 'w', encoding='utf-8') as f:
    f.write("import '../models/quiz_question.dart';\n\n")
    f.write("/// 마태복음 퀴즈 데이터 (29단위 x 5문제 = 145문제) - xlsx에서 자동 생성\n")
    f.write("class MatthewQuizData {\n")
    f.write("  static const Map<int, QuizSet> quizSets = {\n")
    for unit in sorted(quizzes.keys()):
        data = quizzes[unit]
        items = sorted(data['items'], key=lambda x: x[0])
        ch = data['chapter']
        f.write(f"    {unit}: QuizSet(\n")
        f.write(f"      lessonId: 'matthew-lesson-{unit}',\n")
        f.write(f"      title: '마태복음 {ch}장 퀴즈',\n")
        f.write(f"      questions: [\n")
        for idx, (order, q, opts, ci) in enumerate(items):
            opts_str = ', '.join([f"'{esc(o)}'" for o in opts])
            f.write(f"        QuizQuestion(\n")
            f.write(f"          id: 'mq-{unit}-{idx+1}',\n")
            f.write(f"          question: '{q}',\n")
            f.write(f"          options: [{opts_str}],\n")
            f.write(f"          correctIndex: {ci},\n")
            f.write(f"        ),\n")
        f.write(f"      ],\n")
        f.write(f"    ),\n")
    f.write("  };\n\n")
    f.write("  static QuizSet? getQuizForUnit(int unit) => quizSets[unit];\n\n")
    f.write("  static QuizSet getQuizForLesson(String lessonId) {\n")
    f.write("    final match = RegExp(r'matthew-lesson-(\\d+)').firstMatch(lessonId);\n")
    f.write("    if (match != null) {\n")
    f.write("      final unit = int.tryParse(match.group(1)!);\n")
    f.write("      if (unit != null && quizSets.containsKey(unit)) {\n")
    f.write("        return quizSets[unit]!;\n")
    f.write("      }\n")
    f.write("    }\n")
    f.write("    return quizSets.values.first;\n")
    f.write("  }\n")
    f.write("}\n")
print(f"Quiz: {len(quizzes)} units, {sum(len(v['items']) for v in quizzes.values())} questions")


# ===== 3. Meditation data =====
ws3 = wb['묵상 주제_마태복음']
meditations = {}
last_unit = None
for row in ws3.iter_rows(min_row=3, values_only=True):
    topic = row[6]
    if not topic:
        continue

    unit_raw = row[1]
    ch_raw = row[3]
    core = row[4]
    explanation = row[5]
    guide = row[7]

    if unit_raw is not None:
        last_unit = int(float(str(unit_raw)))
    unit = last_unit
    if unit is None:
        continue

    ch = int(float(str(ch_raw))) if ch_raw else None

    if unit not in meditations:
        meditations[unit] = {
            'chapter': ch or 0,
            'core': esc(core) if core else '',
            'explanation': esc(explanation) if explanation else '',
            'topics': []
        }

    meditations[unit]['topics'].append((esc(topic), esc(guide)))

with open(f'{BASE}/study/data/matthew_meditation_data.dart', 'w', encoding='utf-8') as f:
    f.write("import '../models/lesson_data.dart';\n\n")
    f.write("/// 마태복음 묵상 데이터 (29단위, 87개 묵상 주제) - xlsx에서 자동 생성\n")
    f.write("class MatthewMeditationData {\n")

    # LessonContent list
    f.write("  static const List<LessonContent> lessons = [\n")
    for unit in sorted(meditations.keys()):
        m = meditations[unit]
        ch = m['chapter']
        core = m['core']
        explanation = m['explanation'][:300] if m['explanation'] else ''
        first_topic = m['topics'][0][0] if m['topics'] else ''
        title = core if core else f'마태복음 {ch}장'
        scripture = explanation if explanation else f'마태복음 {ch}장을 읽어보세요.'
        f.write(f"    LessonContent(\n")
        f.write(f"      lessonId: 'matthew-lesson-{unit}',\n")
        f.write(f"      pathId: 'path-matthew',\n")
        f.write(f"      title: '{title}',\n")
        f.write(f"      scriptureReference: '마태복음 {ch}장',\n")
        f.write(f"      scriptureText: '{scripture}',\n")
        f.write(f"      meditationGuide: '{first_topic}',\n")
        f.write(f"    ),\n")
    f.write("  ];\n\n")

    # MeditationContent list for meditation screen
    f.write("  /// 묵상 화면용 콘텐츠 (87개)\n")
    f.write("  static const List<MeditationContent> meditationContents = [\n")
    for unit in sorted(meditations.keys()):
        m = meditations[unit]
        ch = m['chapter']
        core = m['core']
        theme = core if core else f'마태복음 {ch}장'
        for topic_text, guide_text in m['topics']:
            f.write(f"    MeditationContent(\n")
            f.write(f"      theme: '{theme}',\n")
            f.write(f"      topic: '{topic_text}',\n")
            f.write(f"      guide: '{guide_text}',\n")
            f.write(f"      chapter: {ch},\n")
            f.write(f"    ),\n")
    f.write("  ];\n")
    f.write("}\n\n")

    f.write("/// 묵상 콘텐츠 모델\n")
    f.write("class MeditationContent {\n")
    f.write("  final String theme;\n")
    f.write("  final String topic;\n")
    f.write("  final String guide;\n")
    f.write("  final int chapter;\n\n")
    f.write("  const MeditationContent({\n")
    f.write("    required this.theme,\n")
    f.write("    required this.topic,\n")
    f.write("    required this.guide,\n")
    f.write("    required this.chapter,\n")
    f.write("  });\n")
    f.write("}\n")

print(f"Meditation: {len(meditations)} units, {sum(len(m['topics']) for m in meditations.values())} topics")
print("Done!")
